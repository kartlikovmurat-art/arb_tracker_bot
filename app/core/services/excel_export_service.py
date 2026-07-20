from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.entities.trade import Trade


class ExcelExportService:

    @staticmethod
    def export(trades: list[Trade]) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Trades"

        headers = [
            "ID",
            "Coin",
            "Buy Exchange",
            "Sell Exchange",
            "Amount",
            "Buy Price",
            "Sell Price",
            "Profit",
            "ROI",
            "Status",
            "Type",
            "Strategy",
            "Created At",
        ]

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column)
            cell.value = header
            cell.font = Font(bold=True)

        for row, trade in enumerate(trades, start=2):

            sheet.cell(row=row, column=1).value = trade.id
            sheet.cell(row=row, column=2).value = trade.coin
            sheet.cell(row=row, column=3).value = trade.buy_exchange
            sheet.cell(row=row, column=4).value = trade.sell_exchange
            sheet.cell(row=row, column=5).value = float(trade.amount)
            sheet.cell(row=row, column=6).value = float(trade.buy_price)
            sheet.cell(row=row, column=7).value = float(trade.sell_price)
            sheet.cell(row=row, column=8).value = float(trade.profit)
            sheet.cell(row=row, column=9).value = float(trade.roi)
            sheet.cell(row=row, column=10).value = trade.status.value
            sheet.cell(row=row, column=11).value = trade.trade_type.value
            sheet.cell(row=row, column=12).value = trade.strategy
            sheet.cell(row=row, column=13).value = trade.created_at.strftime(
                "%Y-%m-%d %H:%M:%S"
            )

        stream = BytesIO()

        workbook.save(stream)

        stream.seek(0)

        return stream